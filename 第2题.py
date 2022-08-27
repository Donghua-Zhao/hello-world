import books as books
import xlrd
import pandas as pd
import numpy as nd
from openpyxl import load_workbook
#books1 = pd.read_excel(r'G:\Desktop\工作簿1.xlsx',sheet_name='题目一4450-4469')
#pingyu = []
#for i in range(0, 20):
#    pingyu.append(books1.iloc[[i], [1]].values[0][0])

#print(pingyu)
#print(type(books1))
#可以通过输出一个数据，检验是否引用正确
#print(books1.iloc[[0], [1]].values[0][0])
work_book = load_workbook(r'G:\Desktop\工作簿1.xlsx')
print(work_book.sheetnames)
sheet_data = work_book['题目一4450-4469']
print(sheet_data)
print(type(sheet_data))
cell_value = sheet_data.cell(row=0, column=0)